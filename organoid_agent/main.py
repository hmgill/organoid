import os
import sys
import asyncio
import traceback
import base64
import argparse
from pathlib import Path

# Ensure the project root is in the Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types

from agents.manager import manager_agent


MIME_TYPE_MAP = {
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png',
    '.webp': 'image/webp',
    '.tif': 'image/tiff',
    '.tiff': 'image/tiff',
}

# MIME types the Gemini API does not accept directly.
# Images with these types are converted to PNG in memory before being sent.
UNSUPPORTED_API_MIME_TYPES = {'image/tiff'}


def encode_image_for_api(image_path: str | Path) -> tuple[str, str]:
    """
    Load an image and return (base64_data, mime_type) ready for the Gemini API.

    TIFF files (and any other formats unsupported by the API) are converted to
    PNG in memory so the API accepts them. The original file on disk is never
    modified, and the original path is still passed to pipeline tools for I/O.

    Args:
        image_path: Path to the source image.

    Returns:
        Tuple of (base64-encoded image data, MIME type string).
    """
    from PIL import Image
    from io import BytesIO

    image_path = Path(image_path)
    mime_type = MIME_TYPE_MAP.get(image_path.suffix.lower(), 'image/jpeg')

    if mime_type in UNSUPPORTED_API_MIME_TYPES:
        print(
            f"  [Encode] {image_path.suffix.upper()} is not supported by the Gemini API "
            f"— converting to PNG in memory (original file unchanged)"
        )
        img = Image.open(image_path).convert("RGB")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        return base64.b64encode(buffer.getvalue()).decode('utf-8'), 'image/png'

    with open(image_path, 'rb') as f:
        return base64.b64encode(f.read()).decode('utf-8'), mime_type


def load_image_paths_from_file(
    file_path: str,
    column: str = None,
    start_row: int = None,
    end_row: int = None,
) -> list[str]:
    """
    Load image paths from a CSV or XLSX file, with optional row range filtering.

    Row numbers are 1-based and refer to data rows (excluding the header).
    For example, start_row=1, end_row=10 loads the first 10 data rows.

    Args:
        file_path:  Path to the CSV or XLSX file.
        column:     Column name containing image paths. Defaults to the first column.
        start_row:  First data row to include (1-based, inclusive). Defaults to 1.
        end_row:    Last data row to include (1-based, inclusive). Defaults to last row.

    Returns:
        List of image path strings.
    """
    suffix = Path(file_path).suffix.lower()

    if suffix == '.csv':
        import csv
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            col = column or reader.fieldnames[0]
            all_paths = [row[col].strip() for row in reader if row[col].strip()]
        print(f"[System] Read {len(all_paths)} total rows from CSV column '{col}'")

    elif suffix in ('.xlsx', '.xls'):
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to read XLSX files. "
                "Install with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_index = headers.index(column) if column else 0
        col_name = headers[col_index]

        all_paths = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_index]
            if val and str(val).strip():
                all_paths.append(str(val).strip())

        wb.close()
        print(f"[System] Read {len(all_paths)} total rows from XLSX column '{col_name}'")

    else:
        raise ValueError(f"Unsupported file format: '{suffix}'. Use .csv or .xlsx")

    # Apply row slice (convert 1-based inclusive range to 0-based Python slice)
    row_start = (start_row - 1) if start_row is not None else 0
    row_end   = end_row if end_row is not None else len(all_paths)
    paths = all_paths[row_start:row_end]

    if start_row is not None or end_row is not None:
        actual_start = row_start + 1
        actual_end   = row_start + len(paths)
        print(
            f"[System] Row filter applied: rows {actual_start}–{actual_end} "
            f"→ {len(paths)} image(s) selected"
        )
    else:
        print(f"[System] Loaded {len(paths)} image path(s) (no row filter)")

    return paths


async def run_workflow(input_path: str, session_index: int = 0):
    """Run the full organoid analysis pipeline for a single image."""

    if not os.path.exists(input_path):
        print(f"[!] ERROR: Image file '{input_path}' not found! Skipping.")
        return False

    print(f"[System] Loading image: {input_path}")
    image_data, mime_type = encode_image_for_api(input_path)
    print(f"[System] Image encoded ({len(image_data)} bytes base64, sending as {mime_type})")

    user_prompt = (
        f"Please analyze this microscopy image of organoids. "
        f"Remove the background and then identify the center coordinates of any organoids you can see. "
        f"The image file is: {input_path}"
    )

    session_service = InMemorySessionService()
    session = await session_service.create_session(
        app_name="organoids",
        user_id="user-1",
        session_id=f"session-{session_index}",
    )

    print(f"[System] Session created: {session.id}")

    runner = Runner(
        app_name="organoids",
        agent=manager_agent,
        session_service=session_service,
    )

    try:
        print("--- Agent Response ---")

        user_message = types.Content(
            role="user",
            parts=[
                types.Part(text=user_prompt),
                types.Part(
                    inline_data=types.Blob(
                        mime_type=mime_type,
                        data=image_data,
                    )
                ),
            ],
        )

        response_text = []
        async for event in runner.run_async(
            user_id=session.user_id,
            session_id=session.id,
            new_message=user_message,
        ):
            if hasattr(event, 'content') and event.content:
                for part in event.content.parts:
                    if hasattr(part, 'text') and part.text:
                        print(part.text, end="", flush=True)
                        response_text.append(part.text)

        print("\n\n[Done]")

        if not response_text:
            print("[!] WARNING: No response text received from the agent")
        else:
            print(f"[Success] Received {len(response_text)} response parts")

        if os.path.exists("outputs"):
            output_files = os.listdir("outputs")
            if output_files:
                print(f"[Success] Output files in outputs/: {output_files}")
            else:
                print("[!] WARNING: outputs/ directory is empty")
        else:
            print("[!] WARNING: outputs/ directory was not created")

        return True

    except Exception as e:
        print(f"\n[!] Runtime Error processing '{input_path}': {e}")
        traceback.print_exc()
        return False


async def run_batch(image_paths: list[str]):
    """Process a list of image paths sequentially."""
    total = len(image_paths)
    succeeded = 0
    failed = 0

    for i, path in enumerate(image_paths, start=1):
        print(f"\n{'='*60}")
        print(f"[Batch] Processing image {i}/{total}: {path}")
        print(f"{'='*60}")
        ok = await run_workflow(path, session_index=i)
        if ok:
            succeeded += 1
        else:
            failed += 1

    print(f"\n{'='*60}")
    print(
        f"[Batch] Complete — {succeeded} succeeded, {failed} failed "
        f"out of {total} images."
    )
    print(f"{'='*60}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Organoid microscopy analysis pipeline. "
            "Accepts a single image path OR a CSV/XLSX file containing image paths."
        )
    )

    input_group = parser.add_mutually_exclusive_group(required=False)
    input_group.add_argument(
        "--image",
        metavar="IMAGE_PATH",
        help="Path to a single input image.",
    )
    input_group.add_argument(
        "--batch",
        metavar="FILE_PATH",
        help="Path to a CSV or XLSX file containing image paths.",
    )

    parser.add_argument(
        "--column",
        metavar="COLUMN_NAME",
        default=None,
        help=(
            "Column name in the CSV/XLSX file that contains image paths. "
            "Defaults to the first column."
        ),
    )

    parser.add_argument(
        "--start-row",
        metavar="N",
        type=int,
        default=None,
        help=(
            "First data row to process, 1-based and inclusive "
            "(e.g. --start-row 1). Defaults to the first row."
        ),
    )

    parser.add_argument(
        "--end-row",
        metavar="N",
        type=int,
        default=None,
        help=(
            "Last data row to process, 1-based and inclusive "
            "(e.g. --end-row 10). Defaults to the last row."
        ),
    )

    # Positional fallback for backwards compatibility (single image, no flag)
    parser.add_argument(
        "image_path",
        nargs="?",
        default=None,
        help="(Legacy) Positional path to a single input image.",
    )

    args = parser.parse_args()

    # Resolve which mode to run
    if args.batch:
        # Batch mode: read paths from CSV/XLSX
        paths = load_image_paths_from_file(
            args.batch,
            column=args.column,
            start_row=args.start_row,
            end_row=args.end_row,
        )
        if not paths:
            print("[!] ERROR: No image paths found in the provided file.")
            sys.exit(1)
        asyncio.run(run_batch(paths))

    else:
        # Single-image mode (--image flag or positional argument)
        single_path = (
            args.image
            or args.image_path
            or "/N/slate/hungill/organoids/data/data/imgs/org64_B2A-2_d19_LabA.jpg"
        )
        asyncio.run(run_workflow(single_path, session_index=1))
